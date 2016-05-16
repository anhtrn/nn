"""conv.py
~~~~~~~~~~

Code for many of the experiments involving convolutional networks in
Chapter 6 of the book 'Neural Networks and Deep Learning', by Michael
Nielsen.  The code essentially duplicates (and parallels) what is in
the text, so this is simply a convenience, and has not been commented
in detail.  Consult the original text for more details.

"""

from collections import Counter

import time
import pandas as pd
from openpyxl import load_workbook
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import theano
import theano.tensor as T

import network3
from network3 import sigmoid, tanh, ReLU, Network
from network3 import ConvPoolLayer, FullyConnectedLayer, SoftmaxLayer

training_data, validation_data, test_data = network3.load_data_shared(
                                                batch1_name="../data/airplane_automobile", 
                                                batch2_name="../data/airplane_automobile_validation",
                                                batch3_name="../data/airplane_automobile_test")
mini_batch_size = 100
n_iterations = 3

def shallow(wbName, n=n_iterations, epochs=30):
    # open output excel file and load to dataframe
    df1 = pd.read_excel(open(wbName, 'rb'), sheetname='Sheet1')    
    book = load_workbook(wbName)
    writer = pd.ExcelWriter(wbName, engine='openpyxl')         
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    nets = []
    for j in range(n):
        print "\nA shallow net with 100 hidden neurons"
        net = Network([
            FullyConnectedLayer(n_in=1024, n_out=100),
            SoftmaxLayer(n_in=100, n_out=2)], mini_batch_size)

        start1 = time.time()
        current_test_accuracy = net.SGD(
            training_data, epochs, mini_batch_size, 0.1, validation_data, test_data)
        end1 = time.time()
        print "*** Time Elapsed %f" % (end1-start1)
        
        # write results to dataframe
        df1.iloc[0, 3*j] = current_test_accuracy
        df1.iloc[0, 3*j+1] = 1 - current_test_accuracy
        df1.iloc[0, 3*j+2] = end1 - start1

        nets.append(net)
        
    # calculate mean and std for all accuracies, errors, and times recorded
    for k in range(0, 3):
        df1.iloc[0, 3*n+2*k] = np.mean([df1.iloc[0, index1] for index1 in range(k, 3*n, 3)])
        df1.iloc[0, 3*n+2*k+1] = np.std([df1.iloc[0, index2] for index2 in range(k, 3*n, 3)])
        
    # save to file
    df1.to_excel(writer, 'Sheet1')
    writer.save()
    
    return nets 

def basic_conv(wbName, n=n_iterations, epochs=30):
    # open output excel file and load to dataframe
    df2 = pd.read_excel(open(wbName, 'rb'), sheetname='Sheet1')    
    book = load_workbook(wbName)
    writer = pd.ExcelWriter(wbName, engine='openpyxl')         
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    nets = []
    for j in range(n):
        print "\nConv + FC architecture"
        net = Network([
            ConvPoolLayer(image_shape=(mini_batch_size, 1, 32, 32), 
                          filter_shape=(20, 1, 5, 5), 
                          poolsize=(2, 2)),
            FullyConnectedLayer(n_in=20*14*14, n_out=100),
            SoftmaxLayer(n_in=100, n_out=2)], mini_batch_size)

        start2 = time.time()
        current_test_accuracy = net.SGD(
            training_data, epochs, mini_batch_size, 0.1, validation_data, test_data)
        end2 = time.time()
        print "*** Time Elapsed %f" % (end2-start2)
        
        # write results to dataframe
        # 3*n+6 is the number of columns already occupied by previous recorded
        # results. For this to work, we are assuming shallow() is run before
        # basic_conv()
        df2.iloc[0, 3*n+6+3*j] = current_test_accuracy
        df2.iloc[0, 3*n+6+3*j+1] = 1 - current_test_accuracy
        df2.iloc[0, 3*n+6+3*j+2] = end2 - start2
        
    # calculate mean and std for all accuracies, errors, and times recorded
    for k in range(0, 3):
        df2.iloc[0, 6*n+6+4*k] = np.mean([df2.iloc[0, index1] for index1 in range(3*n+6+k, 6*n+6, 3)])
        df2.iloc[0, 6*n+6+4*k+1] = np.std([df2.iloc[0, index2] for index2 in range(3*n+6+k, 6*n+6, 3)])
    # calculate additive and multiplicative increase in mean accu and mean error compared to shallow()
    for t in range(0, 2):
        df2.iloc[0, 6*n+6+4*t+2] = (df2.iloc[0, 6*n+6+4*t] - df2.iloc[0, 3*n+2*t])
        df2.iloc[0, 6*n+6+4*t+3] = (df2.iloc[0, 6*n+6+4*t] - df2.iloc[0, 3*n+2*t]) / df2.iloc[0, 3*n+2*t]
        
    # save to file
    df2.to_excel(writer, 'Sheet1')
    writer.save()

    return net
    
'''
def omit_FC():
    for j in range(3):
        print "Conv only, no FC"
        net = Network([
            ConvPoolLayer(image_shape=(mini_batch_size, 1, 28, 28), 
                          filter_shape=(20, 1, 5, 5), 
                          poolsize=(2, 2)),
            SoftmaxLayer(n_in=20*12*12, n_out=10)], mini_batch_size)
        net.SGD(training_data, 60, mini_batch_size, 0.1, validation_data, test_data)
    return net 

def dbl_conv(activation_fn=sigmoid):
    for j in range(3):
        print "Conv + Conv + FC architecture"
        net = Network([
            ConvPoolLayer(image_shape=(mini_batch_size, 1, 28, 28), 
                          filter_shape=(20, 1, 5, 5), 
                          poolsize=(2, 2),
                          activation_fn=activation_fn),
            ConvPoolLayer(image_shape=(mini_batch_size, 20, 12, 12), 
                          filter_shape=(40, 20, 5, 5), 
                          poolsize=(2, 2),
                          activation_fn=activation_fn),
            FullyConnectedLayer(
                n_in=40*4*4, n_out=100, activation_fn=activation_fn),
            SoftmaxLayer(n_in=100, n_out=10)], mini_batch_size)
        net.SGD(training_data, 60, mini_batch_size, 0.1, validation_data, test_data)
    return net 

# The following experiment was eventually omitted from the chapter,
# but I've left it in here, since it's an important negative result:
# basic l2 regularization didn't help much.  The reason (I believe) is
# that using convolutional-pooling layers is already a pretty strong
# regularizer.
def regularized_dbl_conv():
    for lmbda in [0.00001, 0.0001, 0.001, 0.01, 0.1, 1.0, 10.0, 100.0]:
        for j in range(3):
            print "Conv + Conv + FC num %s, with regularization %s" % (j, lmbda)
            net = Network([
                ConvPoolLayer(image_shape=(mini_batch_size, 1, 28, 28), 
                              filter_shape=(20, 1, 5, 5), 
                              poolsize=(2, 2)),
                ConvPoolLayer(image_shape=(mini_batch_size, 20, 12, 12), 
                              filter_shape=(40, 20, 5, 5), 
                              poolsize=(2, 2)),
                FullyConnectedLayer(n_in=40*4*4, n_out=100),
                SoftmaxLayer(n_in=100, n_out=10)], mini_batch_size)
            net.SGD(training_data, 60, mini_batch_size, 0.1, validation_data, test_data, lmbda=lmbda)

def dbl_conv_relu():
    for lmbda in [0.0, 0.00001, 0.0001, 0.001, 0.01, 0.1, 1.0, 10.0, 100.0]:
        for j in range(3):
            print "Conv + Conv + FC num %s, relu, with regularization %s" % (j, lmbda)
            net = Network([
                ConvPoolLayer(image_shape=(mini_batch_size, 1, 28, 28), 
                              filter_shape=(20, 1, 5, 5), 
                              poolsize=(2, 2), 
                              activation_fn=ReLU),
                ConvPoolLayer(image_shape=(mini_batch_size, 20, 12, 12), 
                              filter_shape=(40, 20, 5, 5), 
                              poolsize=(2, 2), 
                              activation_fn=ReLU),
                FullyConnectedLayer(n_in=40*4*4, n_out=100, activation_fn=ReLU),
                SoftmaxLayer(n_in=100, n_out=10)], mini_batch_size)
            net.SGD(training_data, 60, mini_batch_size, 0.03, validation_data, test_data, lmbda=lmbda)

#### Some subsequent functions may make use of the expanded MNIST
#### data.  That can be generated by running expand_mnist.py.

def expanded_data(n=100):
    """n is the number of neurons in the fully-connected layer.  We'll try
    n=100, 300, and 1000.

    """
    expanded_training_data, _, _ = network3.load_data_shared(
        "../data/mnist_expanded.pkl.gz")
    for j in range(3):
        print "Training with expanded data, %s neurons in the FC layer, run num %s" % (n, j)
        net = Network([
            ConvPoolLayer(image_shape=(mini_batch_size, 1, 28, 28), 
                          filter_shape=(20, 1, 5, 5), 
                          poolsize=(2, 2), 
                          activation_fn=ReLU),
            ConvPoolLayer(image_shape=(mini_batch_size, 20, 12, 12), 
                          filter_shape=(40, 20, 5, 5), 
                          poolsize=(2, 2), 
                          activation_fn=ReLU),
            FullyConnectedLayer(n_in=40*4*4, n_out=n, activation_fn=ReLU),
            SoftmaxLayer(n_in=n, n_out=10)], mini_batch_size)
        net.SGD(expanded_training_data, 60, mini_batch_size, 0.03, 
                validation_data, test_data, lmbda=0.1)
    return net 

def expanded_data_double_fc(n=100):
    """n is the number of neurons in both fully-connected layers.  We'll
    try n=100, 300, and 1000.

    """
    expanded_training_data, _, _ = network3.load_data_shared(
        "../data/mnist_expanded.pkl.gz")
    for j in range(3):
        print "Training with expanded data, %s neurons in two FC layers, run num %s" % (n, j)
        net = Network([
            ConvPoolLayer(image_shape=(mini_batch_size, 1, 28, 28), 
                          filter_shape=(20, 1, 5, 5), 
                          poolsize=(2, 2), 
                          activation_fn=ReLU),
            ConvPoolLayer(image_shape=(mini_batch_size, 20, 12, 12), 
                          filter_shape=(40, 20, 5, 5), 
                          poolsize=(2, 2), 
                          activation_fn=ReLU),
            FullyConnectedLayer(n_in=40*4*4, n_out=n, activation_fn=ReLU),
            FullyConnectedLayer(n_in=n, n_out=n, activation_fn=ReLU),
            SoftmaxLayer(n_in=n, n_out=10)], mini_batch_size)
        net.SGD(expanded_training_data, 60, mini_batch_size, 0.03, 
                validation_data, test_data, lmbda=0.1)

def double_fc_dropout(p0, p1, p2, repetitions):
    expanded_training_data, _, _ = network3.load_data_shared(
        "../data/mnist_expanded.pkl.gz")
    nets = []
    for j in range(repetitions):
        print "\n\nTraining using a dropout network with parameters ",p0,p1,p2
        print "Training with expanded data, run num %s" % j
        net = Network([
            ConvPoolLayer(image_shape=(mini_batch_size, 1, 28, 28), 
                          filter_shape=(20, 1, 5, 5), 
                          poolsize=(2, 2), 
                          activation_fn=ReLU),
            ConvPoolLayer(image_shape=(mini_batch_size, 20, 12, 12), 
                          filter_shape=(40, 20, 5, 5), 
                          poolsize=(2, 2), 
                          activation_fn=ReLU),
            FullyConnectedLayer(
                n_in=40*4*4, n_out=1000, activation_fn=ReLU, p_dropout=p0),
            FullyConnectedLayer(
                n_in=1000, n_out=1000, activation_fn=ReLU, p_dropout=p1),
            SoftmaxLayer(n_in=1000, n_out=10, p_dropout=p2)], mini_batch_size)
        net.SGD(expanded_training_data, 40, mini_batch_size, 0.03, 
                validation_data, test_data)
        nets.append(net)
    return nets

def ensemble(nets): 
    """Takes as input a list of nets, and then computes the accuracy on
    the test data when classifications are computed by taking a vote
    amongst the nets.  Returns a tuple containing a list of indices
    for test data which is erroneously classified, and a list of the
    corresponding erroneous predictions.

    Note that this is a quick-and-dirty kluge: it'd be more reusable
    (and faster) to define a Theano function taking the vote.  But
    this works.

    """
    
    test_x, test_y = test_data
    for net in nets:
        i = T.lscalar() # mini-batch index
        net.test_mb_predictions = theano.function(
            [i], net.layers[-1].y_out,
            givens={
                net.x: 
                test_x[i*net.mini_batch_size: (i+1)*net.mini_batch_size]
            })
        net.test_predictions = list(np.concatenate(
            [net.test_mb_predictions(i) for i in xrange(1000)]))
    all_test_predictions = zip(*[net.test_predictions for net in nets])
    def plurality(p): return Counter(p).most_common(1)[0][0]
    plurality_test_predictions = [plurality(p) 
                                  for p in all_test_predictions]
    test_y_eval = test_y.eval()
    error_locations = [j for j in xrange(10000) 
                       if plurality_test_predictions[j] != test_y_eval[j]]
    erroneous_predictions = [plurality(all_test_predictions[j])
                             for j in error_locations]
    print "Accuracy is {:.2%}".format((1-len(error_locations)/10000.0))
    return error_locations, erroneous_predictions

def plot_errors(error_locations, erroneous_predictions=None):
    test_x, test_y = test_data[0].eval(), test_data[1].eval()
    fig = plt.figure()
    error_images = [np.array(test_x[i]).reshape(28, -1) for i in error_locations]
    n = min(40, len(error_locations))
    for j in range(n):
        ax = plt.subplot2grid((5, 8), (j/8, j % 8))
        ax.matshow(error_images[j], cmap = matplotlib.cm.binary)
        ax.text(24, 5, test_y[error_locations[j]])
        if erroneous_predictions:
            ax.text(24, 24, erroneous_predictions[j])
        plt.xticks(np.array([]))
        plt.yticks(np.array([]))
    plt.tight_layout()
    return plt
    
def plot_filters(net, layer, x, y):

    """Plot the filters for net after the (convolutional) layer number
    layer.  They are plotted in x by y format.  So, for example, if we
    have 20 filters after layer 0, then we can call show_filters(net, 0, 5, 4) to
    get a 5 by 4 plot of all filters."""
    filters = net.layers[layer].w.eval()
    fig = plt.figure()
    for j in range(len(filters)):
        ax = fig.add_subplot(y, x, j)
        ax.matshow(filters[j][0], cmap = matplotlib.cm.binary)
        plt.xticks(np.array([]))
        plt.yticks(np.array([]))
    plt.tight_layout()
    return plt
'''

#### Helper method to run all experiments in the book

def run_experiments():

    """Run the experiments described in the book.  Note that the later
    experiments require access to the expanded training data, which
    can be generated by running expand_mnist.py.

    """
    # create an empty Excel spreadsheet
    wbName = 'output.xlsx'
    wb = pd.ExcelWriter(wbName)
    # first row entry (index) is name of current experiment 
    index = ['airplane_automobile']
    # names of each column of results in each run
    column_names = ['accu', 'error', 'time']
    # multiple column names
    mcn = []
    # columns for shallow()
    for i in range(0, n_iterations):
        mcn = mcn + column_names
    mcn = mcn + ['mean accu', 'std accu', 'mean error', 'std error', 'mean time', 'std time']
    # columns for basic_conv()
    for j in range(0, n_iterations):
        mcn = mcn + column_names
    mcn = mcn + ['mean accu', 'std accu', 'additive increase', 'multipli increase', 'mean error', 'std error', 'additive increase', 'multipli increase', 'mean time', 'std time']
    
    # create pandas DataFrame based on information above
    df = pd.DataFrame(index=index, columns=mcn)
    # save to file
    df.to_excel(wb, 'Sheet1')
    wb.save()

    shallow(wbName=wbName)
    basic_conv(wbName=wbName)
    '''
    omit_FC()
    dbl_conv(activation_fn=sigmoid)
    # omitted, but still interesting: regularized_dbl_conv()
    dbl_conv_relu()
    expanded_data(n=100)
    expanded_data(n=300)
    expanded_data(n=1000)
    expanded_data_double_fc(n=100)    
    expanded_data_double_fc(n=300)
    expanded_data_double_fc(n=1000)
    nets = double_fc_dropout(0.5, 0.5, 0.5, 5)
    # plot the erroneous digits in the ensemble of nets just trained
    error_locations, erroneous_predictions = ensemble(nets)
    plt = plot_errors(error_locations, erroneous_predictions)
    plt.savefig("ensemble_errors.png")
    # plot the filters learned by the first of the nets just trained
    plt = plot_filters(nets[0], 0, 5, 4)
    plt.savefig("net_full_layer_0.png")
    plt = plot_filters(nets[0], 1, 8, 5)
    plt.savefig("net_full_layer_1.png")
    '''
